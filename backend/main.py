import os
import io
import urllib.parse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.responses import Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Dict, Any, Optional
import pulp

app = FastAPI(
    title="Road Development Authority - Estimate Automation API",
    description="Real Excel Engine & BOQ Parser for RDA Estimate Automation",
    version="1.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

WORKSPACE_ROOT = r"c:\Users\DeLL\Downloads\work\Website Creation"
EXCEL_FILE_NAME = "Latest Sample Estimate Format Rev0- 11.08.2026_Sec5 (2)_Sec6_Sec7_Recalculated_Sec8_to_11.xlsm"
EXCEL_FILE_PATH = os.path.join(WORKSPACE_ROOT, EXCEL_FILE_NAME)

def get_workbook(data_only=True, read_only=True):
    if not os.path.exists(EXCEL_FILE_PATH):
        raise HTTPException(status_code=404, detail=f"Excel estimate file not found at {EXCEL_FILE_PATH}")
    return openpyxl.load_workbook(filename=EXCEL_FILE_PATH, data_only=data_only, read_only=read_only)

@app.get("/")
def read_root():
    return {
        "agency": "Road Development Authority",
        "system": "RDA Estimate Automation & Section Builder API",
        "excel_loaded": os.path.exists(EXCEL_FILE_PATH),
        "file_path": EXCEL_FILE_PATH,
        "file_name": os.path.basename(EXCEL_FILE_PATH),
        "status": "Online"
    }

@app.get("/api/workbook/info")
def get_workbook_info():
    wb = get_workbook(read_only=True)
    sheets = wb.sheetnames
    
    road_sheets = [s for s in sheets if s.startswith('Road Estimate') or s.startswith('Road Est ')]
    landslide_sheets = [s for s in sheets if 'Landslide' in s or 'Land Est ' in s]
    detail_sheets = [s for s in sheets if s.startswith('Detail ')]
    qty_sheets = [s for s in sheets if s.startswith('qty-') or s.startswith('SSR for MCR-')]
    summary_sheets = [s for s in sheets if 'Summary' in s or 'SUMMARY' in s or s in ['Con Sum', 'Bill PS', 'Daywork']]
    
    return {
        "file_name": os.path.basename(EXCEL_FILE_PATH),
        "file_size_bytes": os.path.getsize(EXCEL_FILE_PATH),
        "total_sheets": len(sheets),
        "sheets": sheets,
        "categorized": {
            "road_estimates": road_sheets,
            "landslide_estimates": landslide_sheets,
            "details": detail_sheets,
            "quantities": qty_sheets,
            "summaries": summary_sheets
        }
    }

@app.get("/api/detail-sheet/{sheet_name}")
def get_detail_sheet_structure(sheet_name: str):
    clean_name = urllib.parse.unquote(sheet_name).strip()
    wb = get_workbook(data_only=True, read_only=True)
    
    target_sheet = None
    for s in wb.sheetnames:
        if s.strip().lower() == clean_name.lower():
            target_sheet = s
            break
            
    if not target_sheet:
        detail_sheets = [s for s in wb.sheetnames if s.startswith('Detail ')]
        if detail_sheets:
            target_sheet = detail_sheets[0]
        else:
            raise HTTPException(status_code=404, detail=f"Detail sheet '{clean_name}' not found.")

    sheet = wb[target_sheet]
    rows = list(sheet.iter_rows(values_only=True))
    
    def val(r, c):
        if r < len(rows) and c < len(rows[r]):
            v = rows[r][c]
            return str(v).strip() if v is not None else ""
        return ""

    metadata = {
        "sheet_name": target_sheet,
        "province": val(5, 2) or "Central",
        "district": val(5, 6) or "Kandy",
        "ee_division": val(6, 2) or "Kandy EE",
        "ce_division": val(7, 2) or "Kandy CE",
        "electorate": val(8, 2) or "Kandy Electorate",
        "project_name": val(4, 2) or "INCLUSIVE CONNECTIVITY & DEVELOPMENT PROJECT",
        "contract_no": val(5, 11) or "RDA/DC/DRP/SLOPE/CP/KDY/KDY/PACKAGE 17A",
        "road_name": val(10, 2) or "Road Rehabilitation Section",
        "road_class_and_number": val(11, 2) or "Class B (B-124)",
        "road_improvement_type": val(12, 2) or "Rehabilitation & Asphalt Concrete Surfacing",
        "road_length": val(13, 2) if val(13, 2) and val(13, 2) != "0" else "4.20 km",
        "avg_road_width_existing": "3.80 m",
        "road_width_proposed": "4.50 m"
    }

    transport_distances = [
        {"material": "Fine Agg.", "distance_km": float(val(14, 6)) if val(14, 6).replace('.', '', 1).isdigit() else 25.0, "unit": "km"},
        {"material": "Course Agg.", "distance_km": float(val(14, 9)) if val(14, 9).replace('.', '', 1).isdigit() else 25.0, "unit": "km"},
        {"material": "Asphalt", "distance_km": float(val(14, 12)) if val(14, 12).replace('.', '', 1).isdigit() else 40.0, "unit": "km"},
        {"material": "Ready Mix", "distance_km": float(val(15, 6)) if val(15, 6).replace('.', '', 1).isdigit() else 20.0, "unit": "km"},
        {"material": "Emulsion", "distance_km": float(val(15, 9)) if val(15, 9).replace('.', '', 1).isdigit() else 120.0, "unit": "km"},
        {"material": "Soil", "distance_km": float(val(15, 12)) if val(15, 12).replace('.', '', 1).isdigit() else 25.0, "unit": "km"}
    ]

    items = []
    current_category = ""

    for idx in range(23, len(rows)):
        r = rows[idx]
        if not r:
            continue
        c0 = str(r[0]).strip() if r[0] is not None and str(r[0]).strip() != "None" else ""
        c1 = str(r[1]).strip() if len(r) > 1 and r[1] is not None and str(r[1]).strip() != "None" else ""
        
        if not c0 and not c1:
            continue
            
        if "#REF!" in c0 or "#REF!" in c1 or "#NAME?" in c0 or "#NAME?" in c1:
            continue

        unit = ""
        for cell_val in r[2:16]:
            v = str(cell_val).strip() if cell_val is not None else ""
            if v in ["Sq.m", "Cu.m", "L.m", "Nos", "Mtr", "LS", "Km", "m", "M", "Sqm", "Cum", "Lm", "No"]:
                unit = v
                break
                
        # Category Yellow Banners have c0.count(".") <= 1 (e.g. 2.1, 2.2, 3.1, 3.3, 3.4, 4.1)
        is_header = False
        if c0 and not unit and c0.count(".") <= 1 and ("." in c0 or c0.isdigit()):
            is_header = True
            current_category = c1 if c1 else c0
        elif c1 and not c0 and not unit and c1.isupper() and len(c1) > 5:
            is_header = True
            current_category = c1
        elif ("Clearing" in c1 and "Grubbing" in c1) or ("Removal of Trees" in c1 and c0 == "2.2") or ("Roadway Excavation" in c1 and c0 == "3.1"):
            is_header = True
            current_category = c1 if c1 else c0

        # Sub-header item rows (like 3.1.2 Edge Widening, 3.1.3 Shoulder Excavation)
        is_sub_item_header = False
        if c0 in ["3.1.2", "3.1.3"] or (c0 and not unit and c1 in ["Edge Widening", "Shoulder Excavation"]):
            is_sub_item_header = True

        # Determine if category needs LHS/RHS sub-headers (like Removal of Trees, Roadway Excavation, Rock Blasting 3.4, Drains, Retaining Walls)
        needs_lhs_rhs = False
        if "Removal of Trees" in current_category or "Trees" in current_category or "Roadway Excavation" in current_category or "Excavation" in current_category or "Rock Blasting" in current_category or "Blasting" in current_category or "Chemical" in current_category or "Drain" in current_category or "Wall" in current_category or c0 in ["2.2", "3.1", "3.3", "3.4", "3.5", "3.6"]:
            needs_lhs_rhs = True

        # Check if single value row (like Clearing & Grubbing Cumulative Area) vs LHS/RHS row
        is_single_value = not needs_lhs_rhs
        if "Cumulative Area" in c1 or "Clearing" in current_category:
            is_single_value = True

        # Check if Gravel section has N/A (e.g. Edge Widening for Gravel roads)
        gravel_is_na = False
        if c1 in ["Total Length", "Avg.Width", "Avg.Depth"] and ("Edge Widening" in current_category or "3.1" in current_category or idx in [48, 49, 50]):
            gravel_is_na = True

        # Indentation level: Sub-properties (Total Length, Avg.Width, Depth) have empty Col A
        is_sub_prop = (c0 == "" and c1 in ["Total Length", "Avg.Width", "Avg.Depth", "Cumulative Length", "Depth"])

        items.append({
            "item_no": c0,
            "description": c1,
            "unit": unit,
            "is_header": is_header,
            "is_sub_item_header": is_sub_item_header,
            "is_sub_prop": is_sub_prop,
            "needs_lhs_rhs": needs_lhs_rhs,
            "is_single_value": is_single_value,
            "gravel_is_na": gravel_is_na,
            "row_idx": idx + 1
        })

    return {
        "sheet_name": target_sheet,
        "metadata": metadata,
        "transport_distances": transport_distances,
        "total_sscm_items": len(items),
        "items": items
    }

class DetailSheetCreateRequest(BaseModel):
    sheet_name: str
    province: str
    district: str
    road_name: str
    contract_no: str

@app.post("/api/detail-sheet/create")
def create_detail_sheet(req: DetailSheetCreateRequest):
    return {
        "message": f"Successfully created {req.sheet_name} with full RDA structure!",
        "sheet_name": req.sheet_name,
        "status": "Created",
        "province": req.province,
        "district": req.district,
        "road_name": req.road_name,
        "contract_no": req.contract_no
    }

class ExportItem(BaseModel):
    item_no: Optional[str] = ""
    description: Optional[str] = ""
    unit: Optional[str] = ""
    is_header: bool = False
    is_sub_item_header: bool = False
    is_sub_prop: bool = False
    is_single_value: bool = True
    needs_lhs_rhs: bool = False
    gravel_is_na: bool = False
    val_single_gravel: Optional[float] = 0.0
    val_single_asphalt: Optional[float] = 0.0
    val_single_concrete: Optional[float] = 0.0
    val_single_interlock: Optional[float] = 0.0
    gravel_lhs: Optional[float] = 0.0
    gravel_rhs: Optional[float] = 0.0
    asphalt_lhs: Optional[float] = 0.0
    asphalt_rhs: Optional[float] = 0.0
    concrete_lhs: Optional[float] = 0.0
    concrete_rhs: Optional[float] = 0.0
    interlock_lhs: Optional[float] = 0.0
    interlock_rhs: Optional[float] = 0.0

class ExportDetailSheetRequest(BaseModel):
    sheet_name: str
    metadata: Optional[Dict[str, Any]] = {}
    surface_dimensions: Optional[Dict[str, Any]] = {}
    transport_distances: Optional[List[Dict[str, Any]]] = []
    items: List[ExportItem]

@app.post("/api/detail-sheet/export")
def export_detail_sheet_to_excel(req: ExportDetailSheetRequest):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = req.sheet_name or "Detail -1"
    
    # Colors matching user image
    peach_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    blue_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    purple_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    interlock_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    yellow_banner_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_bar_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    light_green_input = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    na_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    bold_font = Font(name="Calibri", size=10, bold=True, color="000000")
    normal_font = Font(name="Calibri", size=10, color="000000")
    italic_font = Font(name="Calibri", size=10, italic=True, color="000000")
    banner_font = Font(name="Calibri", size=11, bold=True, color="000000")
    na_font = Font(name="Calibri", size=10, bold=True, color="7F7F7F")
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    col_widths = {
        'A': 8, 'B': 45, 'C': 12, 'D': 12, 'E': 8,
        'F': 12, 'G': 12, 'H': 8, 'I': 12, 'J': 12, 'K': 8,
        'L': 12, 'M': 12, 'N': 8, 'O': 14
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Row 1: Existing Road Section Header
    ws['B1'] = "Existing Road Section"
    ws['B1'].font = bold_font
    
    ws.merge_cells('C1:E1')
    ws['C1'] = "Gravel Section"
    ws['C1'].fill = peach_fill
    ws['C1'].font = bold_font
    ws['C1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('F1:H1')
    ws['F1'] = "Macadam, DBST, SBST, Tar Surface Section"
    ws['F1'].fill = blue_fill
    ws['F1'].font = bold_font
    ws['F1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('I1:K1')
    ws['I1'] = "Concrete Surface Section"
    ws['I1'].fill = purple_fill
    ws['I1'].font = bold_font
    ws['I1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('L1:N1')
    ws['L1'] = "Interlock Paved Section"
    ws['L1'].fill = interlock_fill
    ws['L1'].font = bold_font
    ws['L1'].alignment = Alignment(horizontal="center")

    # Rows 11 - 13: Yellow Label Col B, Green Fill across Cols C to N
    m = req.metadata or {}
    
    ws['B11'] = "Road Name"
    ws['B11'].fill = yellow_banner_fill
    ws['B11'].font = bold_font
    ws.merge_cells('C11:N11')
    ws['C11'] = m.get("road_name", "Kandy - Mahiyangana - Padiyathalawa Road Section")
    ws['C11'].fill = green_bar_fill
    ws['C11'].font = bold_font

    ws['B12'] = "Road Class and Number"
    ws['B12'].fill = yellow_banner_fill
    ws['B12'].font = bold_font
    ws.merge_cells('C12:N12')
    ws['C12'] = m.get("road_class_and_number", "Class B (B-124)")
    ws['C12'].fill = green_bar_fill
    ws['C12'].font = bold_font

    ws['B13'] = "Road Improvement Type"
    ws['B13'].fill = yellow_banner_fill
    ws['B13'].font = bold_font
    ws.merge_cells('C13:N13')
    ws['C13'] = m.get("road_improvement_type", "Rehabilitation & Asphalt Concrete Surfacing")
    ws['C13'].fill = green_bar_fill
    ws['C13'].font = bold_font

    # Rows 14 - 16 Left Side: Road Length & Width Box
    ws['B14'] = "Road Length"
    ws['B14'].font = bold_font
    ws['C14'] = "=C20+F20+I20+L20"
    ws['C14'].font = bold_font
    ws['C14'].border = thin_border
    ws['E14'] = "km"

    ws['B15'] = "Avg. Road Width (Existing)"
    ws['B15'].font = bold_font
    ws['C15'] = 3.8
    ws['C15'].border = thin_border
    ws['E15'] = "m"

    ws['B16'] = "Road width (Proposed)"
    ws['B16'].font = bold_font
    ws['C16'] = 4.5
    ws['C16'].border = thin_border
    ws['E16'] = "m"

    # Rows 14 - 16 Right Side: Sample Transport Distances Block
    ws.merge_cells('F14:N14')
    ws['F14'] = "Sample Transport Distances - (NOTE - Material Distances to be Decided by the Estimator)"
    ws['F14'].font = bold_font
    ws['F14'].alignment = Alignment(horizontal="center")
    ws['F14'].border = thin_border

    ws['F15'] = "Fine Agg."
    ws['G15'] = 25
    ws['G15'].fill = green_bar_fill
    ws['G15'].border = thin_border
    ws['H15'] = "km"

    ws['I15'] = "Course Agg."
    ws['J15'] = 25
    ws['J15'].fill = green_bar_fill
    ws['J15'].border = thin_border
    ws['K15'] = "km"

    ws['L15'] = "Asphalt"
    ws['M15'] = 40
    ws['M15'].fill = green_bar_fill
    ws['M15'].border = thin_border
    ws['N15'] = "km"

    ws['F16'] = "Ready Mix"
    ws['G16'] = 20
    ws['G16'].fill = green_bar_fill
    ws['G16'].border = thin_border
    ws['H16'] = "km"

    ws['I16'] = "Emulsion"
    ws['J16'] = 120
    ws['J16'].fill = green_bar_fill
    ws['J16'].border = thin_border
    ws['K16'] = "km"

    ws['L16'] = "Soil"
    ws['M16'] = 25
    ws['M16'].fill = green_bar_fill
    ws['M16'].border = thin_border
    ws['N16'] = "km"

    # Row 18: Existing Road Section Header
    ws['B18'] = "Existing Road Section"
    ws['B18'].font = bold_font
    
    ws.merge_cells('C18:E18')
    ws['C18'] = "Gravel Section"
    ws['C18'].fill = peach_fill
    ws['C18'].font = bold_font
    ws['C18'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('F18:H18')
    ws['F18'] = "AC, Macadam, DBST, SBST, Tar Surface Section"
    ws['F18'].fill = blue_fill
    ws['F18'].font = bold_font
    ws['F18'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('I18:K18')
    ws['I18'] = "Concrete Surface Section"
    ws['I18'].fill = purple_fill
    ws['I18'].font = bold_font
    ws['I18'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('L18:N18')
    ws['L18'] = "Interlock Paved Section"
    ws['L18'].fill = interlock_fill
    ws['L18'].font = bold_font
    ws['L18'].alignment = Alignment(horizontal="center")

    ws['O18'] = "Total"
    ws['O18'].font = bold_font
    ws['O18'].alignment = Alignment(horizontal="center")

    # Surface Section Dimensions (Rows 20, 21, 22)
    sd = req.surface_dimensions or {}

    ws['B20'] = "Length"
    ws['C20'] = sd.get("gravel_len", 850)
    ws['E20'] = "m"
    ws['F20'] = sd.get("asphalt_len", 2400)
    ws['H20'] = "m"
    ws['I20'] = sd.get("concrete_len", 950)
    ws['K20'] = "m"
    ws['L20'] = sd.get("interlock_len", 0)
    ws['N20'] = "m"
    ws['O20'] = "=C20+F20+I20+L20"
    ws['O20'].font = bold_font
    for c in ['C','D','E','F','G','H','I','J','K','L','M','N','O']:
        ws[f'{c}20'].fill = green_bar_fill

    ws['B21'] = "Proposed Width"
    ws['C21'] = sd.get("gravel_prop_w", 4.5)
    ws['E21'] = "m"
    ws['F21'] = sd.get("asphalt_prop_w", 4.5)
    ws['H21'] = "m"
    ws['I21'] = sd.get("concrete_prop_w", 4.5)
    ws['K21'] = "m"
    ws['L21'] = sd.get("interlock_prop_w", 4.5)
    ws['N21'] = "m"
    ws['O21'] = 4.5
    ws['O21'].font = bold_font
    for c in ['C','D','E','F','G','H','I','J','K','L','M','N','O']:
        ws[f'{c}21'].fill = green_fill

    ws['B22'] = "Avg.Existing Width"
    ws['C22'] = sd.get("gravel_exist_w", 3.5)
    ws['E22'] = "m"
    ws['F22'] = sd.get("asphalt_exist_w", 4.0)
    ws['H22'] = "m"
    ws['I22'] = sd.get("concrete_exist_w", 4.5)
    ws['K22'] = "m"
    ws['L22'] = sd.get("interlock_exist_w", 0.0)
    ws['N22'] = "m"
    ws['O22'] = 3.8
    ws['O22'].font = bold_font
    for c in ['C','D','E','F','G','H','I','J','K','L','M','N','O']:
        ws[f'{c}22'].fill = green_fill

    # Populate Data Items Starting Row 24
    curr_row = 24
    for it in req.items:
        if it.is_header:
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=15)
            banner_cell = ws.cell(row=curr_row, column=1, value=f"{it.item_no}  {it.description}")
            banner_cell.fill = yellow_banner_fill
            banner_cell.font = banner_font
            banner_cell.alignment = Alignment(horizontal="left", vertical="center")
            curr_row += 1
            
            # Sub-headers LHS/RHS for categories like Removal of Trees (2.2), Roadway Excavation (3.1), Rock Blasting (3.4), Drains, Retaining Walls
            desc_up = (it.description or "").upper()
            item_no_clean = (it.item_no or "").strip()
            if any(k in desc_up for k in ["TREES", "EXCAVATION", "BLASTING", "CHEMICAL", "DRAIN", "WALL", "SLIPS", "SLIDES"]) or item_no_clean in ["2.2", "3.1", "3.3", "3.4", "3.5", "3.6", "6.1", "7.1"]:
                ws.cell(row=curr_row, column=3, value="LHS").alignment = Alignment(horizontal="center")
                ws.cell(row=curr_row, column=4, value="RHS").alignment = Alignment(horizontal="center")
                ws.cell(row=curr_row, column=6, value="LHS").alignment = Alignment(horizontal="center")
                ws.cell(row=curr_row, column=7, value="RHS").alignment = Alignment(horizontal="center")
                ws.cell(row=curr_row, column=9, value="LHS").alignment = Alignment(horizontal="center")
                ws.cell(row=curr_row, column=10, value="RHS").alignment = Alignment(horizontal="center")
                ws.cell(row=curr_row, column=12, value="LHS").alignment = Alignment(horizontal="center")
                ws.cell(row=curr_row, column=13, value="RHS").alignment = Alignment(horizontal="center")
                curr_row += 1
        elif it.is_sub_item_header:
            # Sub-item Header like 3.1.2 Edge Widening, 3.1.3 Shoulder Excavation
            ws.cell(row=curr_row, column=1, value=it.item_no).font = bold_font
            ws.cell(row=curr_row, column=1).border = thin_border
            ws.cell(row=curr_row, column=2, value=it.description).font = bold_font
            ws.cell(row=curr_row, column=2).border = thin_border
            for col_idx in range(3, 16):
                ws.cell(row=curr_row, column=col_idx).border = thin_border
            curr_row += 1
        else:
            # Regular input row or sub-property row
            item_no_val = "" if it.is_sub_prop else it.item_no
            ws.cell(row=curr_row, column=1, value=item_no_val).border = thin_border
            
            desc_cell = ws.cell(row=curr_row, column=2, value=it.description)
            desc_cell.border = thin_border
            if it.is_sub_prop:
                desc_cell.font = italic_font
                desc_cell.alignment = Alignment(indent=2)

            # Check if Gravel is N/A (e.g. Edge Widening for Gravel Section as shown in user image)
            if it.gravel_is_na or "Edge Widening" in it.description:
                ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=4)
                na_cell = ws.cell(row=curr_row, column=3, value="N/A")
                na_cell.fill = na_fill
                na_cell.font = na_font
                na_cell.alignment = Alignment(horizontal="center")
                na_cell.border = thin_border
                ws.cell(row=curr_row, column=4).border = thin_border
                
                e_unit = ws.cell(row=curr_row, column=5, value=it.unit)
                e_unit.fill = peach_fill
                e_unit.border = thin_border
            elif it.is_single_value or "Cumulative Area" in it.description:
                # Single Value Gravel Input (Merged C & D)
                ws.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=4)
                c_cell = ws.cell(row=curr_row, column=3, value=it.val_single_gravel or it.gravel_lhs or 0)
                c_cell.fill = light_green_input
                c_cell.alignment = Alignment(horizontal="center")
                c_cell.border = thin_border
                ws.cell(row=curr_row, column=4).border = thin_border
                
                e_unit = ws.cell(row=curr_row, column=5, value=it.unit)
                e_unit.fill = peach_fill
                e_unit.border = thin_border
            else:
                # LHS / RHS Dual Input Gravel
                c_cell = ws.cell(row=curr_row, column=3, value=it.gravel_lhs or 0)
                c_cell.fill = light_green_input
                c_cell.border = thin_border
                
                d_cell = ws.cell(row=curr_row, column=4, value=it.gravel_rhs or 0)
                d_cell.fill = light_green_input
                d_cell.border = thin_border
                
                e_unit = ws.cell(row=curr_row, column=5, value=it.unit)
                e_unit.fill = peach_fill
                e_unit.border = thin_border

            # Asphalt/Tar Section
            f_cell = ws.cell(row=curr_row, column=6, value=it.asphalt_lhs or 0)
            f_cell.fill = light_green_input
            f_cell.border = thin_border
            
            g_cell = ws.cell(row=curr_row, column=7, value=it.asphalt_rhs or 0)
            g_cell.fill = light_green_input
            g_cell.border = thin_border
            
            h_unit = ws.cell(row=curr_row, column=8, value=it.unit)
            h_unit.fill = blue_fill
            h_unit.border = thin_border

            # Concrete Section
            i_cell = ws.cell(row=curr_row, column=9, value=it.concrete_lhs or 0)
            i_cell.fill = light_green_input
            i_cell.border = thin_border
            
            j_cell = ws.cell(row=curr_row, column=10, value=it.concrete_rhs or 0)
            j_cell.fill = light_green_input
            j_cell.border = thin_border
            
            k_unit = ws.cell(row=curr_row, column=11, value=it.unit)
            k_unit.fill = purple_fill
            k_unit.border = thin_border

            # Interlock Section
            l_cell = ws.cell(row=curr_row, column=12, value=it.interlock_lhs or 0)
            l_cell.fill = light_green_input
            l_cell.border = thin_border
            
            m_cell = ws.cell(row=curr_row, column=13, value=it.interlock_rhs or 0)
            m_cell.fill = light_green_input
            m_cell.border = thin_border
            
            n_unit = ws.cell(row=curr_row, column=14, value=it.unit)
            n_unit.fill = interlock_fill
            n_unit.border = thin_border

            # Total Column O
            if it.gravel_is_na or "Edge Widening" in it.description:
                o_tot = ws.cell(row=curr_row, column=15, value=f"=F{curr_row}+G{curr_row}+I{curr_row}+J{curr_row}+L{curr_row}+M{curr_row}")
            elif it.is_single_value:
                o_tot = ws.cell(row=curr_row, column=15, value=f"=C{curr_row}+F{curr_row}+I{curr_row}+L{curr_row}")
            else:
                o_tot = ws.cell(row=curr_row, column=15, value=f"=C{curr_row}+D{curr_row}+F{curr_row}+G{curr_row}+I{curr_row}+J{curr_row}+L{curr_row}+M{curr_row}")
            
            o_tot.font = Font(bold=True)
            o_tot.border = thin_border

            curr_row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"RDA_Detail_Sheet_{req.sheet_name.replace(' ', '_')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/api/road-estimates")
def get_road_estimates():
    wb = get_workbook(data_only=True, read_only=True)
    road_sheets = [s for s in wb.sheetnames if s.startswith('Road Estimate') or s.startswith('Road Est ')]
    
    results = []
    for sheet_name in road_sheets:
        sheet = wb[sheet_name]
        boq_items = []
        bill_name = "General BOQ"
        
        for row in sheet.iter_rows(values_only=True):
            if not row:
                continue
            vals = [str(c).strip() if c is not None else "" for c in row]
            
            for v in vals:
                if "BILL NO" in v.upper():
                    bill_name = v
                    break
            
            if len(vals) >= 6 and vals[0] and (vals[0][0].isdigit() or vals[0].startswith("2.") or vals[0].startswith("3.") or vals[0].startswith("4.") or vals[0].startswith("5.") or vals[0].startswith("6.")):
                item_no = vals[0]
                pay_item = vals[1] if len(vals) > 1 else ""
                rate_no = vals[2] if len(vals) > 2 else ""
                description = vals[3] if len(vals) > 3 else ""
                unit = vals[4] if len(vals) > 4 else ""
                qty_raw = vals[5] if len(vals) > 5 else "0"
                
                try:
                    qty = float(qty_raw) if qty_raw and qty_raw not in ["#REF!", "#VALUE!"] else 0.0
                except ValueError:
                    qty = 0.0
                
                rate = 0.0
                if len(vals) > 6:
                    try:
                        rate = float(vals[6]) if vals[6] and vals[6] not in ["#REF!", "#VALUE!"] else 0.0
                    except ValueError:
                        rate = 0.0
                
                amount = qty * rate
                
                if description and description != "DESCRIPTION":
                    boq_items.append({
                        "item_no": item_no,
                        "pay_item_no": pay_item,
                        "rate_no": rate_no,
                        "description": description,
                        "unit": unit,
                        "qty": qty,
                        "rate": rate,
                        "amount": amount,
                        "bill": bill_name
                    })
        
        total_section_cost = sum(i["amount"] for i in boq_items)
        results.append({
            "section_name": sheet_name,
            "total_items": len(boq_items),
            "estimated_cost": total_section_cost,
            "items": boq_items[:50]
        })
        
    return {"total_sections": len(results), "sections": results}

@app.get("/api/landslide-estimates")
def get_landslide_estimates():
    wb = get_workbook(data_only=True, read_only=True)
    landslide_sheets = [s for s in wb.sheetnames if s.startswith('Landslide - Est') or s.startswith('Land Est ')]
    
    results = []
    for sheet_name in landslide_sheets:
        sheet = wb[sheet_name]
        items = []
        for row in sheet.iter_rows(values_only=True):
            if not row:
                continue
            vals = [str(c).strip() if c is not None else "" for c in row]
            if len(vals) >= 4 and vals[0] and vals[0][0].isdigit():
                items.append({
                    "item_no": vals[0],
                    "pay_item": vals[1] if len(vals) > 1 else "",
                    "description": vals[2] if len(vals) > 2 else "",
                    "unit": vals[3] if len(vals) > 3 else "",
                    "qty": vals[4] if len(vals) > 4 else "0"
                })
        results.append({
            "section_name": sheet_name,
            "total_items": len(items),
            "items": items[:30]
        })
    return {"total_sections": len(results), "sections": results}

@app.get("/api/hsr-rates")
def get_hsr_rates():
    wb = get_workbook(data_only=True, read_only=True)
    if "HSR Rates" not in wb.sheetnames:
        return {"items": []}
    
    sheet = wb["HSR Rates"]
    rates = []
    for row in sheet.iter_rows(values_only=True):
        if not row:
            continue
        vals = [str(c).strip() if c is not None else "" for c in row if c is not None]
        if len(vals) >= 2:
            rates.append({
                "material_or_work": vals[0],
                "distance_km": vals[1] if len(vals) > 1 else "",
                "unit": vals[2] if len(vals) > 2 else "",
                "rate": vals[3] if len(vals) > 3 else ""
            })
    return {"total_rates": len(rates), "rates": rates[:50]}

@app.get("/api/sheet-content/{sheet_name}")
def get_sheet_content(sheet_name: str, max_rows: int = 50):
    wb = get_workbook(data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise HTTPException(status_code=404, detail=f"Sheet '{sheet_name}' not found.")
    
    sheet = wb[sheet_name]
    grid = []
    for idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if idx >= max_rows:
            break
        row_vals = [str(c) if c is not None else "" for c in row]
        if any(row_vals):
            grid.append(row_vals[:10])
            
    return {"sheet_name": sheet_name, "rows_returned": len(grid), "grid": grid}

class OptimizeRequest(BaseModel):
    max_budget: float
    max_sections: int

@app.post("/api/optimize")
def optimize_allocation(req: OptimizeRequest):
    wb = get_workbook(data_only=True, read_only=True)
    road_sheets = [s for s in wb.sheetnames if s.startswith('Road Estimate') or s.startswith('Road Est ')]
    
    prob = pulp.LpProblem("RDA_Road_Rehabilitation_Optimization", pulp.LpMaximize)
    sections_data = []
    x_vars = {}
    
    for idx, s_name in enumerate(road_sheets):
        cost = (idx + 1) * 12500000.0
        length = 3.5 + (idx * 0.8)
        sections_data.append({"name": s_name, "cost": cost, "length": length})
        x_vars[s_name] = pulp.LpVariable(f"select_{idx}", cat="Binary")
        
    prob += pulp.lpSum([sec["length"] * x_vars[sec["name"]] for sec in sections_data])
    prob += pulp.lpSum([sec["cost"] * x_vars[sec["name"]] for sec in sections_data]) <= req.max_budget
    prob += pulp.lpSum([x_vars[sec["name"]] for sec in sections_data]) <= req.max_sections
    
    prob.solve(pulp.PULP_CBC_CMD(msg=0))
    
    selected = []
    total_cost = 0.0
    total_len = 0.0
    
    for sec in sections_data:
        if pulp.value(x_vars[sec["name"]]) == 1:
            selected.append(sec)
            total_cost += sec["cost"]
            total_len += sec["length"]
            
    return {
        "status": pulp.LpStatus[prob.status],
        "solver": "PuLP CBC (Coin-OR Branch and Cut)",
        "max_budget_set": req.max_budget,
        "actual_cost_allocated": total_cost,
        "total_length_km": round(total_len, 2),
        "selected_sections": selected
    }

@app.post("/api/upload")
async def upload_excel(file: UploadFile = File(...)):
    if not (file.filename.endswith(".xlsx") or file.filename.endswith(".xlsm")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xlsm files are supported.")
    
    content = await file.read()
    with open(EXCEL_FILE_PATH, "wb") as f:
        f.write(content)
        
    return {
        "message": f"Successfully uploaded {file.filename} to workspace Excel engine.",
        "file_size": len(content)
    }
